import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from datetime import datetime

def seleccionar_archivo(entry):
    archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    if archivo:
        entry.delete(0, tk.END)
        entry.insert(0, archivo)

def estilos(filename):
    wb = load_workbook(filename)
    ws = wb.activeA

    # Aplicar colores
    orange_fill = PatternFill(start_color='E26B0A', end_color='E26B0A', fill_type='solid')
    for cell in ws['A1:I1']:
        for c in cell:
            c.fill = orange_fill

    blue_fill = PatternFill(start_color='0C0950', end_color='0C0950', fill_type='solid')
    for cell in ws['J1:S1']:
        for c in cell:
            c.fill = blue_fill

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Flexo', size=10)

    # Cambiar la fuente de las cabeceras a blanco y negrita
    for cell in ws[1]:
        cell.font = Font(name='Flexo', size=10, color='FFFFFF', bold=True)

    wb.save(filename)

def aceptar():
    archivo1 = archivo1_entry.get()
    archivo2 = archivo2_entry.get()
    archivo3 = archivo3_entry.get()
    archivo4 = archivo4_entry.get()

    if not archivo1 or not archivo2 or not archivo3 or not archivo4:
        messagebox.showerror("Error", "Por favor, selecciona los tres archivos.")
        return
    archivo1 = archivo1_entry.get()
  
    #-------INPUTS-------------------
    dfRQ = pd.read_excel(archivo1, engine='openpyxl') #rq2
    dfBBAA = pd.read_excel(archivo2, engine='openpyxl')    #BBAA
    dfTP = pd.read_excel(archivo3, engine='openpyxl')    #TP

    dfBuscarMatricula = dfBBAA[['Matrícula', 'Nombres']]
    dfRQ2 = dfRQ[['Cod Req', 'Motivo de requerimiento', 'Moviper/Cese (En caso sea reemplazo)','Estatus de aprobación', 'Área', 'Servicio', 'Matrícula de persona a reemplazar', 'Nombre de persona a reemplazar']]
    dfTP2 = dfTP[['Cod Req F','Job_Opening_Status','Motivo_de_Cancelacion','Indicar_nombre_movimiento_interno','Fecha Ingreso','Documento','AP Paterno','AP Materno','Nombres']]

    # Filtro de Matrícula
    dfRQ2.loc['Estado BBAA'] = dfRQ2['Matrícula de persona a reemplazar'].apply(lambda x: 'Activo' if x in dfBuscarMatricula['Matrícula'].values else 'Cesado')

    #Merge de TP
    merged_df_final = pd.merge(dfRQ2, dfTP2, left_on='Cod Req', right_on='Cod Req F', how='inner')
    merged_df_final = merged_df_final.drop(columns=['Cod Req F'])

    #Añadir campos de control/ Formateo de DF
    merged_df_final['PROYECTAR'] = None
    merged_df_final['ESTADO FINAL'] = None
    merged_df_final.columns = merged_df_final.columns.str.upper()
    merged_df_final.fillna('-', inplace=True)
    merged_df_final = merged_df_final.astype(str)


    
    fecha_actual = datetime.now().strftime('%d%m%y')
    output_filename = f'D:\\Datos de Usuarios\\T42928\\Desktop\\Py Control de Proyectados\\BBDD_Control_Proyectados_{fecha_actual}.xlsx'

    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        merged_df_final.to_excel(writer, index=False)

    print("Se ha creado el archivo con éxito.")
    estilos(output_filename)
    messagebox.showinfo("Éxito", "El archivo se ha generado correctamente.")

# Crear la interfaz gráfica
root = tk.Tk()
root.title("ETL - Ingreso de Archivos")
root.geometry("760x310")
root.configure(bg="#f0f0f0")

# Estilo de los widgets
label_style = {"bg": "#f0f0f0", "font": ("Arial", 12)}
entry_style = {"font": ("Arial", 12)}
button_style = {"bg": "#4CAF50", "fg": "white", "font": ("Arial", 12), "activebackground": "#45a049"}
title_style = {"bg": "#f0f0f0", "font": ("Arial", 16, "bold")}

# Añadir el título
tk.Label(root, text="BASE CONTROL DE PROYECTADOS", **title_style).grid(row=0, column=0, columnspan=3, pady=20)

tk.Label(root, text="Ingresa Excel de Requerimientos:", **label_style).grid(row=2, column=0, padx=10, pady=10, sticky="w")
archivo1_entry = tk.Entry(root, width=40, **entry_style)
archivo1_entry.grid(row=2, column=1, padx=10, pady=10, sticky="w")
archivo1_btn = tk.Button(root, text="Seleccionar", command=lambda: seleccionar_archivo(archivo1_entry), **button_style)
archivo1_btn.grid(row=2, column=2, padx=10, pady=10)

tk.Label(root, text="Ingresa BBAA del corte actual", **label_style).grid(row=3, column=0, padx=10, pady=10, sticky="w")
archivo2_entry = tk.Entry(root, width=40, **entry_style)
archivo2_entry.grid(row=3, column=1, padx=10, pady=10, sticky="w")
archivo2_btn = tk.Button(root, text="Seleccionar", command=lambda: seleccionar_archivo(archivo2_entry), **button_style)
archivo2_btn.grid(row=3, column=2, padx=10, pady=10)

tk.Label(root, text="Ingresa archivo TP:", **label_style).grid(row=4, column=0, padx=10, pady=10, sticky="w")
archivo3_entry = tk.Entry(root, width=40, **entry_style)
archivo3_entry.grid(row=4, column=1, padx=10, pady=10, sticky="w")
archivo3_btn = tk.Button(root, text="Seleccionar", command=lambda: seleccionar_archivo(archivo3_entry), **button_style)
archivo3_btn.grid(row=4, column=2, padx=10, pady=10)

tk.Label(root, text="Ingresa Maestro MOVIPER:", **label_style).grid(row=5, column=0, padx=10, pady=10, sticky="w")
archivo4_entry = tk.Entry(root, width=40, **entry_style)
archivo4_entry.grid(row=5, column=1, padx=10, pady=10, sticky="w")
archivo4_btn = tk.Button(root, text="Seleccionar", command=lambda: seleccionar_archivo(archivo4_entry), **button_style)
archivo4_btn.grid(row=5, column=2, padx=10, pady=10)

aceptar_btn = tk.Button(root, text="Generar Base",command= aceptar, **button_style)
aceptar_btn.grid(row=6, column=1, padx=10, pady=20)

root.mainloop()
